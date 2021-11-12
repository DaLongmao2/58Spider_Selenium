import os
import random
import time
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException


def main():
    driver.get(url='https://cd.58.com/changfang/b5/?PGTID=0d30576d-0001-2875-1747-8a40f302d714&ClickID=5')
    spider(driver)


def spider(driver):
    URL = driver.current_url
    print(URL)
    content = driver.page_source.encode('utf-8')
    Soup = BeautifulSoup(content,'lxml')
    li_list = Soup.find(attrs={'class':'content-side-left'}).find_next(attrs={'class':'list-main-style'}).find_all('li')
    print(len(li_list))

    a_list = [url.find_next('a').get('href') for url in li_list]

    a_list.insert(0, 'https://cd.58.com/fangchan/43991188472711x.shtml?prd=lhyli6AwXxsicGE6LVcokmdTvi8HLOnRB2V6n6UFnBo%3D&houseId=1708830440528905&gpos=67&keyword=&PGTID=0d30576d-01d1-d8bb-cfae-6a92c8183d6b&ClickID=9')
    for a in a_list:
        # try:
        # 生成保存所有数据的列表
        all_list = [f" " for _ in range(33)]
        t = random.randint(1,8)

        if record_read_func(str(a).split('.')[2].split('/')[2]):
            print(f'该网页记录中存在,已被跳过{a}')
            continue
        record_save_func(str(a).split('.')[2].split('/')[2])

        # print(f"本次访问延时{t}秒")
        print(f'正在爬取 {a} 页面中')


        # time.sleep(t)
        driver.get(a)

        # for i in range(10):
            # driver.set_window_size(random.randint(500, 1800), random.randint(100, 600))
        # driver.maximize_window()

        content1 = driver.page_source.encode('utf-8')
        Soup1 = BeautifulSoup(content1, 'lxml')

        # 标题
        div_house_title = Soup1.find(attrs={'class':'house-title'})
        title = div_house_title.find_next('h1').get_text()
        title = str(title).replace('(出售)', '', 1)
        title = ''.join(title.split())
        all_list.insert(0, title)

        # 更新于
        tag_span = div_house_title.find_next('p')
        update_time = tag_span.find_next('span').find_next(attrs={'class': 'up'}).get_text()
        all_list.insert(1, update_time)

        base_info = Soup1.find(attrs={'class':'house-basic-right'})

        # 金额 "价格", "平均每天","是否可面议"
        money = base_info.find_next(attrs={'class':'house_basic_title_money'})
        money_num = money.find_next(attrs={'class':'house_basic_title_money_num'}).get_text()
        if money_num != '面议':
            money_unit = money.find_next(attrs={'class':'house_basic_title_money_unit'}).get_text()
            money_second = money.find_next(attrs={'class':'house_basic_title_money_num_second'}).get_text()
            all_list.insert(2, f"{money_num}{money_unit}")
            mianyi_ = str(money_second).split('（')
            if len(mianyi_) > 1:
                all_list.insert(3, mianyi_[0].strip())
                all_list.insert(4, mianyi_[1].replace('）', '', 1).strip())
        else:
            all_list.insert(2, money_num)

        # 建筑面积  类型 起租面积
        mianji_list = base_info.find_next(attrs={'class':'house_basic_title_info'}).find_all(attrs={'class':'up'})
        mianji = []
        for i in mianji_list:
            mianji.append(i.get_text())
        all_list.insert(5, mianji[0])
        all_list.insert(6, mianji[1])
        all_list.insert(7, mianji[2])

        # 发布人信息
        faburen = base_info.find_next(attrs={'class':'house-basic-poster-wrapper'})

        # 发布人信息链接
        faburen_info_src = faburen.find_next('a').get('href')
        all_list.insert(8, faburen_info_src)

        # 发布人姓名
        faburen_name = faburen.find_next(attrs={'class':'poster-name'}).find_next('span').get_text()
        all_list.insert(9, faburen_name)

        try:
            # 发布所属公司
            faburen_company = faburen.find_next(attrs={'class': 'poster-company-4'}).get_text()
            all_list.insert(10, faburen_company)
        except:
            # 个人
            faburen_company = faburen.find_next('p').get_text()
            all_list.insert(10, faburen_company)

        try:
            # 执照图片
            faburen_zhizhao_src = faburen.find_next(attrs={'class': 'zhizhao'}).find_next('i').find_next('a').get(
                'href')
            all_list.insert(11, faburen_zhizhao_src)

            # 执照编码
            faburen_zhizhao_info = faburen.find_next(attrs={'class': 'zhizhao'}).get_text()
            faburen_zhizhao_info = str(faburen_zhizhao_info).strip()
            all_list.insert(12, faburen_zhizhao_info)

            # 区域
            quyu = base_info.find_next(attrs={"class": "house_basic_title_info_2"}).find_next('p').get_text()
            quyu = str(quyu).split('：')[1].strip()
            all_list.insert(13, quyu)

            # 地址
            dizhi = base_info.find_next(attrs={'class': 'p_2'}).get_text()
            dizhi = str(dizhi).split('：')[1].strip().split('地图')[0].strip()
            all_list.insert(14, dizhi)
        except:
            pass

        # 发布人电话
        try:
            phone_button = wait_time.until(EC.element_to_be_clickable((By.CLASS_NAME, 'house-chat-phone')))
            phone_button.click()
            time.sleep(2)
            content2 = driver.page_source.encode('utf-8')
            Soup2 = BeautifulSoup(content2, 'lxml')
            phone = Soup2.find(attrs={'class': 'phone-after-click'}).get_text()
            if phone == 'TA已开启号码保护，请扫码拨号':
                phone_img_src = Soup2.find(attrs={'id': 'house-chat-phone-bubble'}).find_next('img').get('src')
                all_list.insert(15, f"{phone}{phone_img_src}")
            else:
                all_list.insert(15, phone)
            if not phone:
                print(f'电话获取失败,当前页面为:{a}')
        except:
            print(f'电话获取出错,当前页面为:{a}')


        # ================
        # 信息base
        info_div =  Soup1.find(attrs={'class':'house-detail-left'})

        def pan(list1, key, value, type1, index):
            if key == f'{type1}':
                list1.insert(index, value)

        # 基本信息
        ji_list = info_div.find_next(attrs={'id': 'jichu'}).find_all('li')
        jd_d = {}
        for ji in ji_list:
            jd_d[ji.find_next(attrs={'class': 'title'}).get_text()] = ji.find_next(attrs={'class': 'content'}).get_text()
        for key, value in jd_d.items():
            pan(all_list, key, value, '类型', 16)
            pan(all_list, key, value, '可办环评', 17)
            pan(all_list, key, value, '土地性质', 18)
            pan(all_list, key, value, '产权年限', 19)
            pan(all_list, key, value, '租赁方式', 20)
            pan(all_list, key, value, '地址', 21)

        # 内部详情
        try:
            nei_list = info_div.find_next(attrs={'id': 'neibu'}).find_all('li')
            nei_d = {}
            for nei in nei_list:
                nei_d[nei.find_next(attrs={'class': 'title'}).get_text()] = nei.find_next(
                    attrs={'class': 'content'}).get_text()
            for key, value in nei_d.items():
                pan(all_list, key, value, '楼层', 22)
                pan(all_list, key, value, '首层层高', 23)
                pan(all_list, key, value, '厂房新旧', 24)
                pan(all_list, key, value, '厂房结构', 25)
                pan(all_list, key, value, '楼板承重', 26)
                pan(all_list, key, value, '供电电压', 27)
                pan(all_list, key, value, '厂房特色', 28)
                pan(all_list, key, value, '消防备案', 29)
        except:
            pass

        try:
            # 配套
            pei_list = info_div.find_next(attrs={'class': 'peitao-icon'}).find_all(attrs={'class': 'peitao-on'})
            pei_l = []
            for pei in pei_list:
                pei_l.append(pei.get_text())
            all_list.insert(30, str(pei_l))
        except:
            pass

        # 描述
        miao = info_div.find_next(attrs={'class': 'des-item'}).find_next(attrs={'class':'detail'}).get_text()
        all_list.insert(31, miao)

        # 页面链接
        all_list.insert(32, a)

        # 图片
        tu_list = info_div.find_next(attrs={"id":"generalType"}).find_next('div').find_next('ul').find_all('li')
        tu_l = []
        for tu in tu_list:
            tu_l.append(tu.find_next('img').get('src'))
        all_list.insert(33, str(tu_l))

        global number_cont
        global x

        # driver.get_screenshot_as_file(fr'截图\{faburen_zhizhao_info}.png')
        print(all_list)
        save_file(all_list)

        number_cont += 1
        x += 1
        # driver.forward()
        # break
        # except:
        #     continue
    try:
        driver.get(URL)
        time.sleep(2)
        next_element = wait_time.until(EC.element_to_be_clickable((By.CLASS_NAME,'next')))
        time.sleep(2)
        # 定位到页面的“下一页”的按钮，并进行点击
        next_element.click()
        spider(driver)
        # 进行一个递归的循环
    except (TimeoutException, WebDriverException) as e:
        print("Last page reached")
    except NoSuchElementException:
        print('已经到了末尾了！')
        return


def create_file():
    if check_file in ['Y', 'y']:
        file1 = '截图'

        def creates_files(path):
            isExists = os.path.exists(path)
            if not isExists:
                os.makedirs(path)
                print(f'成功创建 【{path}】 文件')
            else:
                print(f'文件 【{file1}】 已存在')
            return

        creates_files(file1)

        book_create = openpyxl.Workbook()
        sheet = book_create.create_sheet("出售")
        title_name = [
            "标题", "更新于", "价格", "平均每天", "是否可面议", "厂房结构", "建筑面积", "起租面积", "发布人信息", "发布人姓名", "发布所属公司", "营业执照", "执照编码",
            "区域", "地址", "联系方式",
            "类型", "可办环评", "土地性质", "产权年限", "Null", "地址", "楼层", "首层层高", "厂房新旧", "厂房结构", "楼板承重", "供电电压", "厂房特色", "消防备案",
            "配套", "描述", "页面链接", '图片链接']
        for i in range(int(len(title_name))):
            sheet.cell(1, i+1).value = title_name[i]
        book_create.save(u"58同城出售信息.xlsx")
        return
    return


def save_file(data):
    book = openpyxl.load_workbook(u"58同城出售信息.xlsx")
    sheet = book["出售"]

    for i in range(int(len(data))):
        sheet.cell(x, i+1).value = data[i]
    book.save('58同城出售信息.xlsx')
    return


def record_read_func(url):
    """
    存在返回True
    不存在返回False
    :param file:
    :param url:
    :return:
    """
    if record_read not in ['N', 'n']:
        with open(f'{record_read}.txt', 'r') as fp:
            record_data = fp.read()
            record_data_list = record_data.split('\n')
            if url in record_data_list:
                return True
            return False
    return None

def record_save_func(url):
    # 追加
    if record_add:
        with open(f'{record_read}.txt', 'a+') as fp:
            fp.write(f"{url}\n")
        return True

    # 新创建
    if record_name:
        with open(f'{record_name}.txt', 'a+') as fp:
            fp.write(f"{url}\n")
        return True
    return


if __name__ == '__main__':
    number_cont = 1
    x = 2
    record_read = None
    record_add = None
    record_name = None
    while True:
        # y轴 爬取记录读取 爬取记录追加 爬取记录保存 是否创建文件夹
        record_read = input('是否读取爬取记录(读取请直接输入记录文件名称,不读取请输入(N):')
        if record_read not in ['N', 'n']: # 是 y
            try:
                x = int(input('请输入Excel中现在 Y 轴开始的数据行号(第一次爬取,请输入2,慎重！):'))
            except ValueError:
                print('请输入数字')
                continue
            record_add = input(f'本次爬取记录,是否追加记录到 --> 【{record_read}】 文件中(Y/N)：')
            if record_add not in ['Y', 'y']:
                record_name = input('请输入本次爬虫记录名称:')
        else:
            record_name = input('请输入本次爬虫记录名称:')
        check_file = input('是否自动创建文件夹 【截图】 文件 【58同城出售信息.xlsx】(Y/N):')
        check = input('确认信息无误，操作不可逆！(无误请输入Y):')
        create_file()
        if check in ['Y', 'y']:
            break

    options = webdriver.ChromeOptions()
    prefs = {"profile.managed_default_content_settings.images": 2}
    options.add_experimental_option("prefs", prefs)
    options.add_argument('--diable-gpu')
    driver = webdriver.Chrome()

    wait_time = WebDriverWait(driver, 10, 0.5)
    main()